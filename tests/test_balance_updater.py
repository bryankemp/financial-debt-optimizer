"""Tests for balance_updater module."""

import sqlite3
import tempfile
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, Mock, patch

import openpyxl
import pytest

from debt_optimizer.core.balance_updater import (
    BalanceUpdater,
    BalanceUpdaterError,
)


@pytest.fixture
def temp_db():
    """Create temporary Quicken-like SQLite database."""
    with tempfile.NamedTemporaryFile(suffix=".sqlite", delete=False) as f:
        db_path = Path(f.name)

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    # Create tables matching Quicken schema
    cur.execute(
        """
        CREATE TABLE ZACCOUNT (
            Z_PK INTEGER PRIMARY KEY,
            ZNAME TEXT,
            ZTYPENAME TEXT,
            ZACTIVE INTEGER,
            ZONLINEBANKINGLEDGERBALANCEAMOUNT REAL
        )
    """
    )

    cur.execute(
        """
        CREATE TABLE ZTRANSACTION (
            Z_PK INTEGER PRIMARY KEY,
            ZACCOUNT INTEGER,
            ZAMOUNT REAL,
            ZRECONCILESTATUS INTEGER,
            ZDELETIONCOUNT INTEGER,
            ZPOSTEDDATE INTEGER,
            ZENTEREDDATE INTEGER
        )
    """
    )

    # Insert test accounts
    cur.execute(
        """
        INSERT INTO ZACCOUNT (Z_PK, ZNAME, ZTYPENAME, ZACTIVE, ZONLINEBANKINGLEDGERBALANCEAMOUNT)
        VALUES (1, 'PECU Checking', 'CHECKING', 1, 5000.00)
    """
    )

    cur.execute(
        """
        INSERT INTO ZACCOUNT (Z_PK, ZNAME, ZTYPENAME, ZACTIVE, ZONLINEBANKINGLEDGERBALANCEAMOUNT)
        VALUES (2, 'Chase Freedom', 'CREDITCARD', 1, -1500.00)
    """
    )

    cur.execute(
        """
        INSERT INTO ZACCOUNT (Z_PK, ZNAME, ZTYPENAME, ZACTIVE, ZONLINEBANKINGLEDGERBALANCEAMOUNT)
        VALUES (3, 'Ally Savings', 'SAVINGS', 1, 10000.00)
    """
    )

    cur.execute(
        """
        INSERT INTO ZACCOUNT (Z_PK, ZNAME, ZTYPENAME, ZACTIVE, ZONLINEBANKINGLEDGERBALANCEAMOUNT)
        VALUES (4, 'Inactive Account', 'CHECKING', 0, 100.00)
    """
    )

    conn.commit()
    conn.close()

    yield db_path

    db_path.unlink()


@pytest.fixture
def temp_xlsx():
    """Create temporary Excel workbook."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx_path = Path(f.name)

    wb = openpyxl.Workbook()

    # Create Debts sheet
    ws_debts = wb.active
    ws_debts.title = "Debts"
    ws_debts.append(["Account", "Balance", "Rate", "Min Payment"])
    ws_debts.append(["Chase Freedom Card", 1000.00, 18.99, 25])
    ws_debts.append(["Discover IT", 500.00, 15.99, 20])

    # Create Balances sheet
    ws_balances = wb.create_sheet("Balances")
    ws_balances.append(["Type", "Account", "Balance"])
    ws_balances.append(["Checking", "PECU Check", 4500.00])
    ws_balances.append(["Savings", "Ally Save", 9500.00])

    wb.save(xlsx_path)
    wb.close()

    yield xlsx_path

    xlsx_path.unlink()


@pytest.fixture
def mock_fuzz_library():
    """Mock fuzzy matching library."""
    with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
        with patch("debt_optimizer.core.balance_updater.fuzz") as mock_fuzz:
            with patch("debt_optimizer.core.balance_updater.process") as mock_process:
                mock_fuzz.ratio = Mock(return_value=85)
                mock_process.extractOne = Mock(return_value=("Chase Freedom", 85, 0))
                yield mock_fuzz, mock_process


class TestBalanceUpdaterInit:
    """Test BalanceUpdater initialization."""

    def test_init_success(self, temp_db):
        """Test successful initialization."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)
            assert updater.db_path == temp_db
            assert updater.fuzzy_threshold == 80
            assert updater.bank_account_name == "PECU Checking"
            assert updater.auto_backup is True

    def test_init_missing_fuzz_library(self, temp_db):
        """Test initialization fails without fuzzy library."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", False):
            with pytest.raises(ImportError, match="Fuzzy matching library required"):
                BalanceUpdater(temp_db)

    def test_init_db_not_found(self):
        """Test initialization fails with missing database."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            with pytest.raises(FileNotFoundError, match="Quicken database not found"):
                BalanceUpdater(Path("/nonexistent/db.sqlite"))

    def test_init_custom_params(self, temp_db):
        """Test initialization with custom parameters."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(
                temp_db,
                fuzzy_threshold=90,
                bank_account_name="Custom Bank",
                auto_backup=False,
            )
            assert updater.fuzzy_threshold == 90
            assert updater.bank_account_name == "Custom Bank"
            assert updater.auto_backup is False


class TestBalanceUpdaterBackup:
    """Test backup functionality."""

    def test_backup_excel(self, temp_db, temp_xlsx):
        """Test Excel backup creation."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)
            backup_path = updater.backup_excel(temp_xlsx)

            assert backup_path.exists()
            assert ".backup." in backup_path.name
            assert temp_xlsx.stem in backup_path.name

            backup_path.unlink()


class TestBalanceUpdaterDatabase:
    """Test database operations."""

    def test_connect_db(self, temp_db):
        """Test database connection."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)
            conn = updater.connect_db()

            assert isinstance(conn, sqlite3.Connection)
            conn.close()

    def test_load_quicken_balances(self, temp_db):
        """Test loading balances from Quicken database."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)
            accounts, credit_cards, checking, savings = updater.load_quicken_balances()

            # Verify account data
            assert "PECU Checking" in accounts
            assert "Chase Freedom" in accounts
            assert "Ally Savings" in accounts
            assert "Inactive Account" not in accounts  # Inactive accounts excluded

            # Verify balances
            assert accounts["PECU Checking"]["balance"] == 5000.00
            assert accounts["Chase Freedom"]["balance"] == -1500.00
            assert accounts["Ally Savings"]["balance"] == 10000.00

            # Verify account types
            assert accounts["PECU Checking"]["type"] == "CHECKING"
            assert accounts["Chase Freedom"]["type"] == "CREDITCARD"
            assert accounts["Ally Savings"]["type"] == "SAVINGS"

            # Verify categorization
            assert "Chase Freedom" in credit_cards
            assert "PECU Checking" in checking
            assert "Ally Savings" in savings


class TestBalanceUpdaterPrompt:
    """Test user prompt functionality."""

    def test_prompt_yes_no_default_no(self, temp_db):
        """Test prompt with default no."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            with patch("builtins.input", return_value=""):
                result = updater._prompt_yes_no("Continue?", default_no=True)
                assert result is False

    def test_prompt_yes_no_default_yes(self, temp_db):
        """Test prompt with default yes."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            with patch("builtins.input", return_value=""):
                result = updater._prompt_yes_no("Continue?", default_no=False)
                assert result is True

    def test_prompt_yes_no_explicit_yes(self, temp_db):
        """Test prompt with explicit yes."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            for answer in ["y", "Y", "yes", "YES"]:
                with patch("builtins.input", return_value=answer):
                    result = updater._prompt_yes_no("Continue?")
                    assert result is True

    def test_prompt_yes_no_explicit_no(self, temp_db):
        """Test prompt with explicit no."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            for answer in ["n", "N", "no", "NO"]:
                with patch("builtins.input", return_value=answer):
                    result = updater._prompt_yes_no("Continue?")
                    assert result is False

    def test_prompt_yes_no_eof(self, temp_db):
        """Test prompt with EOF."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            with patch("builtins.input", side_effect=EOFError):
                result = updater._prompt_yes_no("Continue?", default_no=True)
                assert result is False


class TestBalanceUpdaterUpdateDebts:
    """Test debt update functionality."""

    def test_update_debts_exact_match(self, temp_db, temp_xlsx):
        """Test updating debts with exact match."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            wb = openpyxl.load_workbook(temp_xlsx)
            ws = wb["Debts"]
            ws.cell(row=2, column=1).value = "Chase Freedom"
            ws.cell(row=2, column=2).value = 1000.00

            accounts = {
                "Chase Freedom": {"id": 1, "type": "CREDITCARD", "balance": -1500.00}
            }
            credit_cards = ["Chase Freedom"]

            updates = updater.update_debts_sheet(ws, accounts, credit_cards)

            assert len(updates) == 1
            assert updates[0]["auto"] is True
            assert updates[0]["score"] == 100
            assert updates[0]["new_balance"] == 1500.00
            assert ws.cell(row=2, column=2).value == 1500.00

    def test_update_debts_no_change(self, temp_db, temp_xlsx):
        """Test updating debts when balance hasn't changed."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            wb = openpyxl.load_workbook(temp_xlsx)
            ws = wb["Debts"]
            ws.cell(row=2, column=1).value = "Chase Freedom"
            ws.cell(row=2, column=2).value = 1500.00

            accounts = {
                "Chase Freedom": {"id": 1, "type": "CREDITCARD", "balance": -1500.00}
            }
            credit_cards = ["Chase Freedom"]

            updates = updater.update_debts_sheet(ws, accounts, credit_cards)

            assert len(updates) == 0

    def test_update_debts_empty_sheet(self, temp_db):
        """Test updating empty debts sheet."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Account", "Balance"])

            accounts = {}
            credit_cards = []

            updates = updater.update_debts_sheet(ws, accounts, credit_cards)

            assert len(updates) == 0

    def test_update_debts_no_credit_cards(self, temp_db, temp_xlsx):
        """Test updating debts with no credit cards."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            wb = openpyxl.load_workbook(temp_xlsx)
            ws = wb["Debts"]

            accounts = {}
            credit_cards = []

            updates = updater.update_debts_sheet(ws, accounts, credit_cards)

            assert len(updates) == 0

    def test_update_debts_fuzzy_match_approved(self, temp_db, temp_xlsx):
        """Test fuzzy match with user approval."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            with patch(
                "debt_optimizer.core.balance_updater.process.extractOne"
            ) as mock_extract:
                mock_extract.return_value = ("Chase Freedom", 85, 0)

                updater = BalanceUpdater(temp_db)

                wb = openpyxl.load_workbook(temp_xlsx)
                ws = wb["Debts"]
                # Clear existing data and add only one row
                ws.delete_rows(2, ws.max_row)
                ws.append(["Chase Card", 1000.00, 18.99, 25])

                accounts = {
                    "Chase Freedom": {
                        "id": 1,
                        "type": "CREDITCARD",
                        "balance": -1500.00,
                    }
                }
                credit_cards = ["Chase Freedom"]

                with patch.object(updater, "_prompt_yes_no", return_value=True):
                    updates = updater.update_debts_sheet(ws, accounts, credit_cards)

                assert len(updates) == 1
                assert updates[0]["auto"] is False
                assert updates[0]["score"] == 85
                assert ws.cell(row=2, column=1).value == "Chase Freedom"

    def test_update_debts_fuzzy_match_rejected(self, temp_db, temp_xlsx):
        """Test fuzzy match with user rejection."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            with patch(
                "debt_optimizer.core.balance_updater.process.extractOne"
            ) as mock_extract:
                mock_extract.return_value = ("Chase Freedom", 85, 0)

                updater = BalanceUpdater(temp_db)

                wb = openpyxl.load_workbook(temp_xlsx)
                ws = wb["Debts"]

                accounts = {
                    "Chase Freedom": {
                        "id": 1,
                        "type": "CREDITCARD",
                        "balance": -1500.00,
                    }
                }
                credit_cards = ["Chase Freedom"]

                with patch.object(updater, "_prompt_yes_no", return_value=False):
                    updates = updater.update_debts_sheet(ws, accounts, credit_cards)

                assert len(updates) == 0

    def test_update_debts_score_too_low(self, temp_db, temp_xlsx):
        """Test fuzzy match below threshold."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            with patch(
                "debt_optimizer.core.balance_updater.process.extractOne"
            ) as mock_extract:
                mock_extract.return_value = ("Chase Freedom", 50, 0)

                updater = BalanceUpdater(temp_db)

                wb = openpyxl.load_workbook(temp_xlsx)
                ws = wb["Debts"]

                accounts = {
                    "Chase Freedom": {
                        "id": 1,
                        "type": "CREDITCARD",
                        "balance": -1500.00,
                    }
                }
                credit_cards = ["Chase Freedom"]

                updates = updater.update_debts_sheet(ws, accounts, credit_cards)

                assert len(updates) == 0

    def test_update_debts_rapidfuzz_format(self, temp_db, temp_xlsx):
        """Test handling rapidfuzz 2-tuple return format."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            with patch(
                "debt_optimizer.core.balance_updater.process.extractOne"
            ) as mock_extract:
                # rapidfuzz returns 2-tuple
                mock_extract.return_value = ("Chase Freedom", 85)

                updater = BalanceUpdater(temp_db)

                wb = openpyxl.load_workbook(temp_xlsx)
                ws = wb["Debts"]
                # Clear existing data and add only one row
                ws.delete_rows(2, ws.max_row)
                ws.append(["Chase Card", 1000.00, 18.99, 25])

                accounts = {
                    "Chase Freedom": {
                        "id": 1,
                        "type": "CREDITCARD",
                        "balance": -1500.00,
                    }
                }
                credit_cards = ["Chase Freedom"]

                with patch.object(updater, "_prompt_yes_no", return_value=True):
                    updates = updater.update_debts_sheet(ws, accounts, credit_cards)

                assert len(updates) == 1


class TestBalanceUpdaterUpdateSettings:
    """Test settings update functionality."""

    def test_update_settings_exact_match(self, temp_db, temp_xlsx):
        """Test updating settings with exact match."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            wb = openpyxl.load_workbook(temp_xlsx)
            if "Settings" not in wb.sheetnames:
                ws = wb.create_sheet("Settings")
            else:
                ws = wb["Settings"]
            ws.cell(row=3, column=2).value = 4000.00

            accounts = {
                "PECU Checking": {"id": 1, "type": "CHECKING", "balance": 5000.00}
            }
            checking = ["PECU Checking"]

            result = updater.update_settings_sheet(ws, accounts, checking)

            assert result is not None
            assert result["matched"] == "exact"
            assert result["balance"] == 5000.00
            assert ws.cell(row=3, column=2).value == 5000.00

    def test_update_settings_no_change(self, temp_db, temp_xlsx):
        """Test settings update when balance unchanged."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            wb = openpyxl.load_workbook(temp_xlsx)
            if "Settings" not in wb.sheetnames:
                ws = wb.create_sheet("Settings")
            else:
                ws = wb["Settings"]
            ws.cell(row=3, column=2).value = 5000.00

            accounts = {
                "PECU Checking": {"id": 1, "type": "CHECKING", "balance": 5000.00}
            }
            checking = ["PECU Checking"]

            result = updater.update_settings_sheet(ws, accounts, checking)

            assert result is None

    def test_update_settings_fuzzy_match_approved(self, temp_db, temp_xlsx):
        """Test fuzzy match in settings with approval."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            with patch(
                "debt_optimizer.core.balance_updater.process.extractOne"
            ) as mock_extract:
                mock_extract.return_value = ("PECU Checking Account", 90, 0)

                updater = BalanceUpdater(temp_db, bank_account_name="PECU Check")

                wb = openpyxl.load_workbook(temp_xlsx)
                if "Settings" not in wb.sheetnames:
                    ws = wb.create_sheet("Settings")
                else:
                    ws = wb["Settings"]
                ws.cell(row=3, column=2).value = 4000.00

                accounts = {
                    "PECU Checking Account": {
                        "id": 1,
                        "type": "CHECKING",
                        "balance": 5000.00,
                    }
                }
                checking = ["PECU Checking Account"]

                with patch.object(updater, "_prompt_yes_no", return_value=True):
                    result = updater.update_settings_sheet(ws, accounts, checking)

                assert result is not None
                assert "fuzzy" in result["matched"]
                assert result["balance"] == 5000.00

    def test_update_settings_fuzzy_match_rejected(self, temp_db, temp_xlsx):
        """Test fuzzy match in settings with rejection."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            with patch(
                "debt_optimizer.core.balance_updater.process.extractOne"
            ) as mock_extract:
                mock_extract.return_value = ("PECU Checking Account", 90, 0)

                updater = BalanceUpdater(temp_db, bank_account_name="PECU Check")

                wb = openpyxl.load_workbook(temp_xlsx)
                if "Settings" not in wb.sheetnames:
                    ws = wb.create_sheet("Settings")
                else:
                    ws = wb["Settings"]

                accounts = {
                    "PECU Checking Account": {
                        "id": 1,
                        "type": "CHECKING",
                        "balance": 5000.00,
                    }
                }
                checking = ["PECU Checking Account"]

                with patch.object(updater, "_prompt_yes_no", return_value=False):
                    result = updater.update_settings_sheet(ws, accounts, checking)

                assert result is None

    def test_update_settings_no_checking_accounts(self, temp_db, temp_xlsx):
        """Test settings update with no checking accounts."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            wb = openpyxl.load_workbook(temp_xlsx)
            if "Settings" not in wb.sheetnames:
                ws = wb.create_sheet("Settings")
            else:
                ws = wb["Settings"]

            accounts = {}
            checking = []

            result = updater.update_settings_sheet(ws, accounts, checking)

            assert result is None


class TestBalanceUpdaterUpdateWorkbook:
    """Test full workbook update."""

    def test_update_workbook_success(self, temp_db, temp_xlsx):
        """Test successful workbook update."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            # Setup workbook
            wb = openpyxl.load_workbook(temp_xlsx)
            if "Settings" not in wb.sheetnames:
                wb.create_sheet("Settings")
            # Update debt to exact match to avoid prompts
            ws = wb["Debts"]
            ws.cell(row=2, column=1).value = "Chase Freedom"
            wb.save(temp_xlsx)
            wb.close()

            with patch.object(updater, "load_quicken_balances") as mock_load:
                mock_load.return_value = (
                    {
                        "Chase Freedom": {
                            "id": 1,
                            "type": "CREDITCARD",
                            "balance": -1500.00,
                        }
                    },
                    ["Chase Freedom"],
                    [],
                    [],
                )

                result = updater.update_workbook(temp_xlsx, interactive=False)

                assert result["workbook_path"] == temp_xlsx
                assert result["backup_path"] is not None
                assert isinstance(result["debt_updates"], list)

    def test_update_workbook_file_not_found(self, temp_db):
        """Test workbook update with missing file."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            with pytest.raises(FileNotFoundError, match="Excel workbook not found"):
                updater.update_workbook(Path("/nonexistent/file.xlsx"))

    def test_update_workbook_no_debts_sheet(self, temp_db, temp_xlsx):
        """Test workbook update without Debts sheet."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db)

            # Create workbook without Debts sheet
            wb = openpyxl.Workbook()
            wb.save(temp_xlsx)
            wb.close()

            with patch.object(updater, "load_quicken_balances") as mock_load:
                mock_load.return_value = ({}, [], [], [])

                with pytest.raises(
                    BalanceUpdaterError, match="'Debts' sheet not found"
                ):
                    updater.update_workbook(temp_xlsx)

    def test_update_workbook_no_backup(self, temp_db, temp_xlsx):
        """Test workbook update without auto backup."""
        with patch("debt_optimizer.core.balance_updater.HAS_FUZZ", True):
            updater = BalanceUpdater(temp_db, auto_backup=False)

            # Setup workbook
            wb = openpyxl.load_workbook(temp_xlsx)
            if "Settings" not in wb.sheetnames:
                wb.create_sheet("Settings")
            wb.save(temp_xlsx)
            wb.close()

            with patch.object(updater, "load_quicken_balances") as mock_load:
                mock_load.return_value = ({}, [], [], [])

                result = updater.update_workbook(temp_xlsx)

                assert result["backup_path"] is None


class TestBalanceUpdaterError:
    """Test error handling."""

    def test_balance_updater_error(self):
        """Test custom exception."""
        error = BalanceUpdaterError("Test error")
        assert str(error) == "Test error"
        assert isinstance(error, Exception)
