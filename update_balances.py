#!/usr/bin/env python3
import argparse
import os
import sys
import sqlite3
import time
from datetime import datetime
from pathlib import Path
import shutil

try:
    from thefuzz import process, fuzz  # pip install "thefuzz[speedup]"
except Exception:
    try:
        from rapidfuzz import process, fuzz  # pip install rapidfuzz
    except Exception:
        print("Error: Missing fuzzy matching library.\nInstall one of:\n  pip3 install \"thefuzz[speedup]\"\n  or\n  pip3 install rapidfuzz")
        sys.exit(1)

import openpyxl


def backup_excel(xlsx_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = xlsx_path.with_name(f"{xlsx_path.name}.backup.{timestamp}")
    shutil.copy2(xlsx_path, backup_path)
    print(f"[backup] Created backup: {backup_path}")
    return backup_path


def connect_db(db_path: Path) -> sqlite3.Connection:
    if not db_path.exists():
        print(f"Error: Quicken DB not found at {db_path}")
        sys.exit(1)
    # Open read-only
    return sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)


def load_quicken_balances(conn: sqlite3.Connection):
    """
    Load account balances from Quicken database.
    
    Prefers ZONLINEBANKINGLEDGERBALANCEAMOUNT when available, otherwise sums
    transactions up to the current date. Uses Apple Cocoa timestamp format
    (seconds since 2001-01-01) for date comparisons.
    """
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    # Calculate current Cocoa timestamp (seconds since 2001-01-01)
    cocoa_now = int(time.time() - 978307200)
    
    sql = """
        SELECT
            a.Z_PK AS id,
            a.ZNAME AS name,
            a.ZTYPENAME AS type,
            a.ZACTIVE AS active,
            COALESCE(
                a.ZONLINEBANKINGLEDGERBALANCEAMOUNT,
                SUM(CASE 
                    WHEN t.ZPOSTEDDATE IS NOT NULL AND t.ZPOSTEDDATE <= ?
                    THEN t.ZAMOUNT 
                    ELSE 0 
                END),
                0
            ) AS balance
        FROM ZACCOUNT a
        LEFT JOIN ZTRANSACTION t ON t.ZACCOUNT = a.Z_PK
        WHERE a.ZACTIVE = 1
          AND a.ZTYPENAME IN ('CREDITCARD','CHECKING','SAVINGS')
        GROUP BY a.Z_PK, a.ZNAME, a.ZTYPENAME, a.ZACTIVE, a.ZONLINEBANKINGLEDGERBALANCEAMOUNT
    """
    cur.execute(sql, (cocoa_now,))
    rows = cur.fetchall()

    accounts_by_name = {}
    credit_card_names = []
    checking_names = []
    savings_names = []

    for r in rows:
        name = r["name"]
        acc_type = (r["type"] or "").upper().strip()
        bal = float(r["balance"] or 0.0)
        accounts_by_name[name] = {
            "id": r["id"],
            "type": acc_type,
            "balance": bal,
        }
        if acc_type == "CREDITCARD":
            credit_card_names.append(name)
        elif acc_type == "CHECKING":
            checking_names.append(name)
        elif acc_type == "SAVINGS":
            savings_names.append(name)

    return accounts_by_name, credit_card_names, checking_names, savings_names


def prompt_yes_no(question: str, default_no=True) -> bool:
    default = "n" if default_no else "y"
    prompt = f"{question} [{'Y/n' if not default_no else 'y/N'}]: "
    try:
        ans = input(prompt).strip().lower()
    except EOFError:
        ans = ""
    if not ans:
        ans = default
    return ans in ("y", "yes")


def update_debts_sheet(ws, accounts_by_name, credit_card_names, threshold: int):
    updates = []
    if ws.max_row < 2:
        print("[debts] No data rows found.")
        return updates

    if not credit_card_names:
        print("[debts] No active credit card accounts found in Quicken.")
        return updates

    # Iterate rows starting from row 2
    for row in range(2, ws.max_row + 1):
        excel_name_cell = ws.cell(row=row, column=1)  # Column A
        balance_cell = ws.cell(row=row, column=2)     # Column B

        excel_name = (excel_name_cell.value or "").strip()
        if not excel_name:
            continue

        old_balance = balance_cell.value

        # If exact match among CCs, update without prompt
        if excel_name in credit_card_names:
            qname = excel_name
            qb = accounts_by_name[qname]["balance"]
            new_balance = abs(qb)
            balance_cell.value = float(new_balance)
            # Align name exactly (in case of whitespace differences)
            excel_name_cell.value = qname
            updates.append({
                "row": row,
                "excel_name_old": excel_name,
                "excel_name_new": qname,
                "old_balance": old_balance,
                "new_balance": new_balance,
                "score": 100,
                "auto": True,
            })
            print(f"[debts] Row {row}: exact match '{qname}' -> balance set to {new_balance}")
            continue

        # Fuzzy match to credit card names
        match = process.extractOne(excel_name, credit_card_names, scorer=fuzz.WRatio)
        if not match:
            print(f"[debts] Row {row}: No match candidates for '{excel_name}'. Skipping.")
            continue

        # Handle both rapidfuzz (2-tuple) and thefuzz (3-tuple) return formats
        if len(match) == 3:
            candidate, score, _ = match
        else:
            candidate, score = match
        if score < threshold:
            print(f"[debts] Row {row}: Best match for '{excel_name}' is '{candidate}' (score {score}) below threshold {threshold}. Skipping.")
            continue

        print("\n[debts] Potential match found:")
        print(f"  Excel name   : {excel_name}")
        print(f"  Quicken name : {candidate}")
        print(f"  Score        : {score}")
        if prompt_yes_no("Approve this match?", default_no=True):
            qb = accounts_by_name[candidate]["balance"]
            new_balance = abs(qb)
            balance_cell.value = float(new_balance)
            # Update name to canonical Quicken name so future runs won't prompt again
            excel_name_cell.value = candidate
            updates.append({
                "row": row,
                "excel_name_old": excel_name,
                "excel_name_new": candidate,
                "old_balance": old_balance,
                "new_balance": new_balance,
                "score": score,
                "auto": False,
            })
            print(f"[debts] Row {row}: Updated '{excel_name}' -> '{candidate}', balance {new_balance}")
        else:
            print(f"[debts] Row {row}: Skipped by user for '{excel_name}'")

    return updates


def update_settings_sheet(ws, accounts_by_name, checking_names):
    # Find PECU Checking exactly
    target_name = "PECU Checking"
    if target_name in accounts_by_name and accounts_by_name[target_name]["type"] == "CHECKING":
        bal = float(accounts_by_name[target_name]["balance"])
        ws.cell(row=3, column=2).value = bal
        print(f"[settings] Set Current Bank Balance (B3) from '{target_name}' to {bal}")
        return {"name": target_name, "balance": bal, "matched": "exact"}

    # If exact not found, try fuzzy among checking
    if checking_names:
        match = process.extractOne(target_name, checking_names, scorer=fuzz.WRatio)
        if match:
            # Handle both rapidfuzz (2-tuple) and thefuzz (3-tuple) return formats
            if len(match) == 3:
                candidate, score, _ = match
            else:
                candidate, score = match
            print("\n[settings] Could not find exact 'PECU Checking'.")
            print(f"  Closest checking account: {candidate} (score {score})")
            if prompt_yes_no("Use this account for Current Bank Balance?", default_no=True):
                bal = float(accounts_by_name[candidate]["balance"])
                ws.cell(row=3, column=2).value = bal
                print(f"[settings] Set Current Bank Balance (B3) from '{candidate}' to {bal}")
                return {"name": candidate, "balance": bal, "matched": f"fuzzy:{score}"}

    print("[settings] Warning: 'PECU Checking' not found; Settings!B3 not updated.")
    return None


def main():
    default_db = Path(os.path.expanduser("~/Documents/Bryan.quicken/data"))
    # Prefer the directory that actually has the workbook
    candidate_dirs = [
        Path("/Users/bryan/Projects/Financial"),
        Path("/Users/bryan/Projects/Finances"),
    ]
    xlsx_path = None
    for d in candidate_dirs:
        p = d / "default.xlsx"
        if p.exists():
            xlsx_path = p
            break
    if xlsx_path is None:
        # Fall back to the specified default
        xlsx_path = Path("/Users/bryan/Projects/Financial/default.xlsx")

    parser = argparse.ArgumentParser(description="Update Excel balances from Quicken database with interactive fuzzy matching.")
    parser.add_argument("--db", default=str(default_db), help="Path to Quicken SQLite database (default: %(default)s)")
    parser.add_argument("--xlsx", default=str(xlsx_path), help="Path to Excel workbook (default: %(default)s)")
    parser.add_argument("--threshold", type=int, default=80, help="Fuzzy match threshold (default: %(default)s)")
    args = parser.parse_args()

    db_path = Path(args.db)
    xlsx_path = Path(args.xlsx)

    if not xlsx_path.exists():
        print(f"Error: Excel workbook not found at {xlsx_path}")
        sys.exit(1)

    # Connect DB and load balances
    conn = connect_db(db_path)
    try:
        accounts_by_name, cc_names, checking_names, _ = load_quicken_balances(conn)
    finally:
        conn.close()

    # Load workbook
    wb = openpyxl.load_workbook(str(xlsx_path))

    # Create backup before any changes
    backup_excel(xlsx_path)

    # Get sheets
    if "Debts" not in wb.sheetnames:
        print("Error: 'Debts' sheet not found in workbook.")
        sys.exit(1)
    ws_debts = wb["Debts"]

    if "Settings" not in wb.sheetnames:
        print("Warning: 'Settings' sheet not found; skipping Settings update.")
        ws_settings = None
    else:
        ws_settings = wb["Settings"]

    # Update Debts (credit cards)
    debt_updates = update_debts_sheet(ws_debts, accounts_by_name, cc_names, args.threshold)

    # Update Settings (PECU Checking)
    settings_update = None
    if ws_settings is not None:
        settings_update = update_settings_sheet(ws_settings, accounts_by_name, checking_names)

    # Save workbook
    wb.save(str(xlsx_path))
    print(f"[save] Workbook saved: {xlsx_path}")

    # Summary
    print("\nSummary:")
    if debt_updates:
        print("  Debts updates:")
        for u in debt_updates:
            auto = "auto" if u["auto"] else "approved"
            print(f"    Row {u['row']}: {u['excel_name_old']} -> {u['excel_name_new']} | {u['old_balance']} -> {u['new_balance']} ({auto}, score {u['score']})")
    else:
        print("  Debts updates: none")

    if settings_update:
        print(f"  Settings: Current Bank Balance set from '{settings_update['name']}' to {settings_update['balance']} ({settings_update['matched']})")
    else:
        print("  Settings: no change")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nAborted by user.")
        sys.exit(130)
